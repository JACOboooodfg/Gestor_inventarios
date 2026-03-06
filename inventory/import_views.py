from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db import transaction
from openpyxl import load_workbook
import json
import re
import os
import uuid
import tempfile
from .models import Category, Location, Article
from .forms import ImportExcelForm


# MAPEO DE CÓDIGOS DE UBICACIÓN
LOCATION_CODES = {
    'RECP': 'Recepción',
    'SC': 'Secretaría',
    'ECON': 'Economía',
    'CONT': 'Contadora',
    'AUXCONT': 'Aux. Contabilidad',
    'RECT': 'Rectoría',
    'SD': 'Sala de Docentes',
    'SDC': 'Cafetería Docentes',
    'LABIOG': 'Laboratorio de Biología',
    'LABFIS': 'Laboratorio de Física',
    'FIS': 'Herramientas de Física',
    'DP': 'Deportes',
    'ALM': 'Almacén',
    'PS': 'Psicopedagogía',
    'CA': 'Coordinación Académica',
    'ORT': 'Oratorio',
    'SACR': 'Sacristía',
    'CC': 'Coordinación de Convivencia',
    'LABQUI': 'Laboratorio de Química',
    'SERVG': 'Cuarto Servicios Generales',
    'TEAT': 'Teatro',
    'RT': 'Restaurante/Cafetería',
    'SJ': 'Salón de Juegos',
    'PREK': 'Prekinder',
    'KIN': 'Kinder',
    'JD': 'Jardín',
    'TR': 'Transición',
    'PR': 'Primero',
    'SG': 'Segundo',
    'BT': 'Biblioteca',
    'SLT': 'Sala de Lectura',
    'ARCHCONT': 'Archivo Contabilidad',
    'TDIPR': 'Sala TDI Primaria',
    'SISBTO': 'Sistemas Bachillerato',
    'ROB': 'Robótica',
    'MTOEQUI': 'Cuarto Mto Equipos',
    'SIP': 'Sistemas Primaria',
    'TC': 'Tercero',
    'CT': 'Cuarto',
    'QT': 'Quinto',
    'SX': 'Sexto',
    'SP1': '701',
    'EF': 'Enfermería',
    'TEC': 'Técnicas',
    'PAST': 'Pastoral',
    'DIBUJ': 'Salón de Dibujo',
    'SP2': '702',
    'OC1': '801',
    'OC2': '802',
    'NV1': '901',
    'LBIN': 'Laboratorio de Inglés',
    'NV2': '902',
    'DC1': '1001',
    'DC2': '1002',
    'ONC1': '1101',
    'ONC2': '1102',
    'BAN': 'Cuarto de Banda',
    'TDIB': 'Sala TDI Bachillerato',
    'SAUD': 'Audiovisuales',
    'DZ': 'Danzas',
    'MUS': 'Música',
    'MTO': 'Cuarto de Mantenimiento',
    'PAP': 'Útiles, Papelería',
    'VEHIC': 'Vehículos',
    'DC': 'Portátiles Docentes',
    'CS': 'Cámaras de Seguridad',
    'SUCL4': 'Sillas Universitarias',
    'COM10': 'Elementos Muebles y Enseres',
    'COM02': 'Elementos Equipos y Máquinas',
}


# ─────────────────────────────────────────────
#  HELPERS (sin cambios de lógica)
# ─────────────────────────────────────────────

def infer_category_from_description(description, location_name=''):
    if not description:
        return 'General'
    desc_lower = str(description).lower()
    loc_lower = str(location_name).lower() if location_name else ''
    loc_upper = str(location_name).upper() if location_name else ''

    if 'BT' in loc_upper: return 'Biblioteca'
    if 'PAP' in loc_upper: return 'Papelería'
    if 'DP' in loc_upper or 'DEP' in loc_upper: return 'Deportes'
    if 'LAB' in loc_upper: return 'Ciencias'
    if 'MUS' in loc_upper or 'BAN' in loc_upper: return 'Música'
    if any(c in loc_upper for c in ['TDI', 'SIST', 'ROBOT', 'AV']): return 'Tecnología'
    if any(c in loc_upper for c in ['CAF', 'COC', 'REST', 'COMED']): return 'Cocina y Cafetería'

    if any(w in loc_lower for w in ['biblioteca', 'bibliot', 'lectura']): return 'Biblioteca'
    if any(w in loc_lower for w in ['deport', 'gimnasio', 'cancha']): return 'Deportes'
    if any(w in loc_lower for w in ['papeler', 'utiles', 'útiles']): return 'Papelería'
    if any(w in loc_lower for w in ['laboratorio', 'lab', 'biolog', 'quimic', 'fisica']): return 'Ciencias'
    if any(w in loc_lower for w in ['sistemas', 'informatica', 'audiovisual']): return 'Tecnología'
    if any(w in loc_lower for w in ['music', 'banda']): return 'Música'
    if any(w in loc_lower for w in ['cafeter', 'cocina', 'restaurante', 'comedor']): return 'Cocina y Cafetería'
    if any(w in loc_lower for w in ['secretar', 'recep', 'rector', 'admin', 'oficina']): return 'Oficina'
    if any(w in loc_lower for w in ['aseo', 'limpieza', 'servicio']): return 'Aseo y Limpieza'
    if any(w in loc_lower for w in ['teatro', 'danza', 'arte']): return 'Arte y Teatro'
    if any(w in loc_lower for w in ['orator', 'sacr', 'capilla', 'iglesia']): return 'Elementos de Culto'

    construccion_words = [
        'roseta', 'interruptor', 'bombillo', 'bombilla', 'toma', 'enchufe',
        'cable', 'alambre', 'switch electrico', 'breaker', 'fusible',
        'vidrio', 'ventana', 'puerta', 'cerradura', 'chapa', 'bisagra',
        'tornillo', 'tuerca', 'clavo', 'puntilla', 'cemento', 'arena',
        'pintura pared', 'brocha', 'rodillo pintar', 'lija',
        'tuberia', 'tubería', 'llave agua', 'grifo', 'sifón', 'sifon',
        'codo', 'unión', 'pegante pvc', 'cinta teflón', 'cinta teflon'
    ]
    if any(w in desc_lower for w in construccion_words): return 'Construcción y Mantenimiento'

    papeleria_words = [
        'papel', 'hoja', 'cuaderno', 'lapiz', 'lápiz', 'esfero', 'marcador',
        'resaltador', 'borrador', 'tijera', 'grapadora', 'perforadora', 'carpeta',
        'folder', 'cinta', 'pegante', 'silicona', 'cartulina', 'pintura tempera',
        'colores', 'tempera', 'pincel', 'mina', 'sacapunta', 'regla', 'compás',
        'clips', 'corrector', 'block', 'acetato', 'tinta'
    ]
    if any(w in desc_lower for w in papeleria_words): return 'Papelería'

    deportes_words = [
        'balón', 'balon', 'pelota', 'red', 'cancha', 'aro', 'inflador', 'bomba aire',
        'conos', 'lazo', 'colchoneta', 'cajon', 'baston', 'raqueta', 'guantes deporte',
        'casco', 'futbol', 'fútbol', 'basquet', 'voleibol', 'tenis', 'ping pong',
        'ajedrez', 'uniforme deportivo', 'silbato'
    ]
    if any(w in desc_lower for w in deportes_words): return 'Deportes'

    tech_words = [
        'computador', 'laptop', 'pc', 'monitor', 'teclado', 'mouse', 'impresora',
        'proyector', 'tablet', 'disco duro', 'router', 'switch', 'amplificador',
        'parlante', 'micrófono', 'microfono', 'camara', 'cámara', 'transformador',
        'dvr', 'pantalla', 'cpu', 'scanner', 'usb', 'video beam', 'bateria',
        'cargador', 'hdmi', 'vga'
    ]
    if any(w in desc_lower for w in tech_words): return 'Tecnología'

    ciencias_words = [
        'microscopio', 'probeta', 'beaker', 'tubo ensayo', 'pipeta', 'reactivo',
        'matraz', 'bureta', 'mechero', 'balanza', 'erlenmeyer', 'gradilla',
        'pinza laboratorio'
    ]
    if any(w in desc_lower for w in ciencias_words): return 'Ciencias'

    muebles_words = [
        'silla', 'mesa', 'escritorio', 'estante', 'anaquel', 'archivo', 'armario',
        'locker', 'gabinete', 'pupitre', 'banco', 'sofá', 'sofa', 'butaco', 'mueble',
        'tapete', 'alfombra', 'cajonera', 'repisa', 'perchero', 'vitrina'
    ]
    if any(w in desc_lower for w in muebles_words): return 'Muebles y Enseres'

    cocina_words = [
        'cafetera', 'dispensador', 'nevera', 'estufa', 'horno', 'olla', 'plato',
        'vaso', 'taza', 'cuchara', 'tenedor', 'cuchillo', 'bandeja', 'purificador',
        'licuadora', 'microondas'
    ]
    if any(w in desc_lower for w in cocina_words): return 'Cocina y Cafetería'

    musica_words = [
        'instrumento', 'guitarra', 'piano', 'batería', 'bateria', 'flauta', 'tambor',
        'bombo', 'platillo', 'baqueta', 'trompeta', 'clarinete', 'saxofon', 'violin'
    ]
    if any(w in desc_lower for w in musica_words): return 'Música'

    culto_words = [
        'imagen religiosa', 'virgen', 'santo', 'santa', 'cruz', 'crucifijo',
        'sagrado', 'sagrario', 'altar', 'vela', 'rosario'
    ]
    if any(w in desc_lower for w in culto_words): return 'Elementos de Culto'

    aseo_words = [
        'escoba', 'trapero', 'recogedor', 'balde', 'caneca', 'detergente', 'jabón',
        'jabon', 'trapeador', 'cepillo limpieza', 'guantes aseo', 'limpiador',
        'desinfectante', 'cloro'
    ]
    if any(w in desc_lower for w in aseo_words): return 'Aseo y Limpieza'

    teatro_words = [
        'telón', 'telon', 'cortina escenario', 'escenario', 'tramoya', 'pendon',
        'vestuario', 'disfraz', 'maquillaje teatro'
    ]
    if any(w in desc_lower for w in teatro_words): return 'Arte y Teatro'

    oficina_words = [
        'sello', 'tampón', 'tampon', 'numerador', 'calculadora', 'telefono',
        'teléfono', 'fax', 'archivador oficina'
    ]
    if any(w in desc_lower for w in oficina_words): return 'Oficina'

    return 'General'


def detect_status_from_columns(row_data, headers):
    bueno_col = regular_col = malo_col = None
    for idx, header in enumerate(headers):
        h_clean = str(header).lower().strip()
        if h_clean in ['bueno', 'estado_bueno']:
            bueno_col = idx
        elif h_clean in ['regular', 'estado_regular']:
            regular_col = idx
        elif h_clean in ['malo', 'estado_malo']:
            malo_col = idx

    if bueno_col is not None and len(row_data) > bueno_col:
        if str(row_data[bueno_col]).strip().upper() == 'X':
            return 'available'
    if regular_col is not None and len(row_data) > regular_col:
        if str(row_data[regular_col]).strip().upper() == 'X':
            return 'maintenance'
    if malo_col is not None and len(row_data) > malo_col:
        if str(row_data[malo_col]).strip().upper() == 'X':
            return 'damaged'
    return 'available'


def infer_category_from_sheet_name(sheet_name):
    sheet_lower = str(sheet_name).lower()
    if any(w in sheet_lower for w in ['comunidad', 'listado', 'codificacion', 'codificación', 'rosario', 'funza']):
        return None
    if 'papeler' in sheet_lower or 'útiles' in sheet_lower or 'utiles' in sheet_lower: return 'Papelería'
    if 'equipo' in sheet_lower or 'máquina' in sheet_lower or 'maquina' in sheet_lower: return 'Tecnología'
    if 'deport' in sheet_lower or 'recre' in sheet_lower: return 'Deportes'
    if 'laboratorio' in sheet_lower: return 'Ciencias'
    if 'comun' in sheet_lower or 'radio' in sheet_lower: return 'Tecnología'
    if 'cocina' in sheet_lower or 'cafeter' in sheet_lower: return 'Cocina y Cafetería'
    if 'culto' in sheet_lower: return 'Elementos de Culto'
    if 'music' in sheet_lower or 'instrum' in sheet_lower: return 'Música'
    if 'bibliot' in sheet_lower or 'audiov' in sheet_lower or 'medios' in sheet_lower: return 'Biblioteca'
    if 'mueble' in sheet_lower or 'ensere' in sheet_lower: return 'Muebles y Enseres'
    if 'aseo' in sheet_lower or 'limpieza' in sheet_lower: return 'Aseo y Limpieza'
    if 'vehiculo' in sheet_lower or 'vehículo' in sheet_lower or 'herramient' in sheet_lower: return 'Construcción y Mantenimiento'
    if 'enfermer' in sheet_lower: return 'Enfermería'
    if 'uniforme' in sheet_lower or 'vestuario' in sheet_lower: return 'Vestuario'
    if 'vario' in sheet_lower: return 'General'
    return 'General'


# ─────────────────────────────────────────────
#  TEMP FILE HELPERS  ← NUEVO
#  Guarda el Excel en disco en vez de la sesión
#  para evitar el OOM al serializar en base64.
# ─────────────────────────────────────────────

TEMP_UPLOAD_DIR = tempfile.gettempdir()  # /tmp en Linux


def _save_temp_excel(uploaded_file):
    """Guarda el InMemoryUploadedFile en disco y devuelve la ruta."""
    file_id = str(uuid.uuid4())
    tmp_path = os.path.join(TEMP_UPLOAD_DIR, f'import_{file_id}.xlsx')
    with open(tmp_path, 'wb') as f:
        for chunk in uploaded_file.chunks(chunk_size=1024 * 1024):  # 1 MB chunks
            f.write(chunk)
    return tmp_path, file_id


def _get_temp_excel_path(file_id):
    return os.path.join(TEMP_UPLOAD_DIR, f'import_{file_id}.xlsx')


def _delete_temp_excel(file_id):
    try:
        os.remove(_get_temp_excel_path(file_id))
    except OSError:
        pass


# ─────────────────────────────────────────────
#  VIEWS
# ─────────────────────────────────────────────

@login_required
def import_preview(request):
    """Vista previa para archivo Excel del colegio con múltiples hojas."""

    if request.method == 'POST':

        # ── PASO 1: subida del archivo ──────────────────────────────────────
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']

            try:
                # 🔑 CAMBIO CLAVE: guardar en disco, NO en sesión
                tmp_path, file_id = _save_temp_excel(excel_file)
                request.session['excel_file_id'] = file_id
                request.session['excel_filename'] = excel_file.name

                sheets_info = []

                # Leer solo headers + 5 filas de preview (read_only ahorra RAM)
                wb = load_workbook(tmp_path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]

                    headers = []
                    for cell in ws[8]:
                        if cell.value:
                            headers.append(str(cell.value).strip())
                    if not headers:
                        continue

                    preview_data = []
                    for row_idx, row in enumerate(
                        ws.iter_rows(min_row=9, max_row=13, values_only=True), start=9
                    ):
                        if any(row):
                            preview_data.append({
                                'row_number': row_idx,
                                'data': [str(v) if v is not None else '' for v in row],
                            })

                    sheets_info.append({
                        'name': sheet_name,
                        'headers': headers,
                        'preview': preview_data,
                        'row_count': ws.max_row - 8,
                    })
                wb.close()

                

                return render(request, 'inventory/import_preview_colegio.html', {
                    'sheets_info': sheets_info,
                    'total_sheets': len(sheets_info),
                    'step': 'preview',
                })

            except Exception as e:
                messages.error(request, f'Error al leer el archivo: {str(e)}')
                return redirect('import_preview')

        # ── PASO 2: confirmación de importación ────────────────────────────
        elif 'confirm_import' in request.POST:
            file_id = request.session.get('excel_file_id')
            if not file_id:
                messages.error(request, 'Sesión expirada. Por favor, sube el archivo nuevamente.')
                return redirect('import_preview')

            tmp_path = _get_temp_excel_path(file_id)
            if not os.path.exists(tmp_path):
                messages.error(request, 'El archivo temporal ya no existe. Por favor, sube el archivo nuevamente.')
                return redirect('import_preview')

            try:
                result = import_colegio_excel(tmp_path, request.user)

                if result['success_count'] > 0:
                    messages.success(request, f'✓ {result["success_count"]} artículos importados exitosamente')
                if result['warning_count'] > 0:
                    messages.warning(request, f'⚠ {result["warning_count"]} advertencias')
                if result['error_count'] > 0:
                    messages.error(request, f'✗ {result["error_count"]} errores')
                    for error in result['errors'][:10]:
                        messages.error(request, error)
                    if len(result['errors']) > 10:
                        messages.info(request, f'... y {len(result["errors"]) - 10} errores más')

            except Exception as e:
                messages.error(request, f'Error en la importación: {str(e)}')
            finally:
                # Limpiar siempre: disco + sesión
                _delete_temp_excel(file_id)
                for key in ['excel_file_id', 'excel_filename', 'sheets_info']:
                    request.session.pop(key, None)

            return redirect('article_list')

    return render(request, 'inventory/import_preview_colegio.html', {'step': 'upload'})


# ─────────────────────────────────────────────
#  IMPORTACIÓN CON BATCH  ← OPTIMIZADO
# ─────────────────────────────────────────────

BATCH_SIZE = 50  # Registros por bulk_create (ajustable)


def import_colegio_excel(file_path, user):
    """
    Importar Excel del colegio con múltiples hojas.
    Lee en modo read_only e inserta en batches para minimizar uso de RAM.
    """
    from datetime import datetime

    wb = load_workbook(file_path, read_only=True, data_only=True)

    results = {
        'success_count': 0,
        'error_count': 0,
        'warning_count': 0,
        'errors': [],
        'warnings': [],
    }

    sheet_number = 0
    codigo_counts = {}

    # ── Pre-cargar categorías y ubicaciones existentes en dicts ──────────
    # Evita N consultas a la BD; solo hace GET o INSERT cuando es necesario.
    category_cache = {}   # name_lower → Category instance
    location_cache = {}   # name_lower → Location instance

    def get_or_create_category(name):
        key = name.lower()
        if key not in category_cache:
            obj, _ = Category.objects.get_or_create(
                name__iexact=name, defaults={'name': name}
            )
            category_cache[key] = obj
        return category_cache[key]

    def get_or_create_location(name):
        key = name.lower()
        if key not in location_cache:
            obj, _ = Location.objects.get_or_create(
                name__iexact=name, defaults={'name': name}
            )
            location_cache[key] = obj
        return location_cache[key]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        category_name = infer_category_from_sheet_name(sheet_name)
        if category_name is None:
            results['warnings'].append(f"Hoja '{sheet_name}' ignorada")
            results['warning_count'] += 1
            continue

        sheet_number += 1
        category = get_or_create_category(category_name)

        # Leer encabezados (fila 8)
        headers = []
        for cell in ws[8]:
            if cell.value:
                headers.append(str(cell.value).strip())
        if not headers:
            continue

        # Detectar índices de columnas
        codigo_idx    = next((i for i, h in enumerate(headers) if 'codigo'      in h.lower()), None)
        desc_idx      = next((i for i, h in enumerate(headers) if 'descripcion' in h.lower()), None)
        fecha_idx     = next((i for i, h in enumerate(headers) if 'fecha'       in h.lower() and 'compra' in h.lower()), None)
        cantidad_idx  = next((i for i, h in enumerate(headers) if 'cantidad'    in h.lower()), None)
        valor_idx     = next((i for i, h in enumerate(headers) if 'valor'       in h.lower()), None)
        proveedor_idx = 8
        if any('proveedor' in str(h).lower() or 'comprado' in str(h).lower() or 'lugar' in str(h).lower() for h in headers):
            proveedor_idx = next(
                (i for i, h in enumerate(headers)
                 if h and ('proveedor' in str(h).lower() or 'comprado' in str(h).lower() or 'lugar' in str(h).lower())),
                8,
            )

        # ── Procesar filas en batches ─────────────────────────────────────
        batch = []

        for row_num, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
            try:
                if not any(row):
                    continue

                codigo_original = str(row[codigo_idx]).strip() if codigo_idx is not None and len(row) > codigo_idx and row[codigo_idx] else ''
                descripcion     = str(row[desc_idx]).strip()   if desc_idx   is not None and len(row) > desc_idx   and row[desc_idx]   else ''

                if not descripcion or descripcion.lower() in ['none', 'null', '']:
                    continue

                # Código único
                if codigo_original:
                    codigo_counts[codigo_original] = codigo_counts.get(codigo_original, 0) + 1
                    codigo_unico = f"{codigo_original}-{codigo_counts[codigo_original]:04d}"
                else:
                    codigo_unico = f"AUTO-{sheet_number:02d}-{row_num:04d}"

                # Cantidad
                cantidad = 1
                if cantidad_idx is not None and len(row) > cantidad_idx and row[cantidad_idx]:
                    try:
                        cantidad = int(float(str(row[cantidad_idx]).replace(',', '').strip()))
                    except Exception:
                        cantidad = 1

                # Precio
                precio = None
                if valor_idx is not None and len(row) > valor_idx and row[valor_idx]:
                    try:
                        precio_str = str(row[valor_idx]).replace('$', '').replace(',', '').replace('.', '').strip()
                        if precio_str and precio_str != 'None':
                            precio = float(precio_str)
                    except Exception:
                        pass

                # Fecha de compra
                fecha_compra = None
                if fecha_idx is not None and len(row) > fecha_idx and row[fecha_idx]:
                    try:
                        fecha_val = row[fecha_idx]
                        if isinstance(fecha_val, str):
                            for fmt in ['%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y']:
                                try:
                                    fecha_compra = datetime.strptime(fecha_val.strip(), fmt).date()
                                    break
                                except Exception:
                                    continue
                        elif isinstance(fecha_val, datetime):
                            fecha_compra = fecha_val.date()
                        elif hasattr(fecha_val, 'date'):
                            fecha_compra = fecha_val.date() if callable(fecha_val.date) else fecha_val
                    except Exception:
                        pass

                # Proveedor
                proveedor = None
                if proveedor_idx is not None and len(row) > proveedor_idx and row[proveedor_idx]:
                    prov_val = str(row[proveedor_idx]).strip()
                    if prov_val and prov_val.lower() not in ['none', 'null', '', 'n/a', '-']:
                        proveedor = prov_val[:200]

                # Ubicación
                location_name = 'No Especificado'
                location_code = codigo_original.split('.')[0] if '.' in codigo_original else ''
                if location_code and location_code in LOCATION_CODES:
                    location_name = LOCATION_CODES[location_code]
                location = get_or_create_location(location_name)

                # Estado
                status = detect_status_from_columns(row, headers)

                # Construir objeto (sin guardar todavía)
                article = Article(
                    name=descripcion[:200],
                    code=codigo_unico[:50],
                    category=category,
                    location=location,
                    quantity=max(0, cantidad),
                    unit='unidad',
                    status=status,
                    created_by=user,
                    min_quantity=1,
                )
                if precio:
                    article.price = precio
                if fecha_compra:
                    article.purchase_date = fecha_compra
                if proveedor:
                    article.supplier = proveedor

                batch.append(article)

                # 🔑 BATCH INSERT: cada BATCH_SIZE filas, volcar a la BD
                if len(batch) >= BATCH_SIZE:
                    with transaction.atomic():
                        Article.objects.bulk_create(batch, ignore_conflicts=False)
                    results['success_count'] += len(batch)
                    batch.clear()

            except Exception as e:
                results['errors'].append(f"Hoja '{sheet_name}' Fila {row_num}: {str(e)}")
                results['error_count'] += 1

        # Volcar el último batch de la hoja
        if batch:
            try:
                with transaction.atomic():
                    Article.objects.bulk_create(batch, ignore_conflicts=False)
                results['success_count'] += len(batch)
                batch.clear()
            except Exception as e:
                results['errors'].append(f"Hoja '{sheet_name}' - Error en batch final: {str(e)}")
                results['error_count'] += len(batch)

    wb.close()
    return results


# ─────────────────────────────────────────────
#  OTRAS VISTAS (sin cambios)
# ─────────────────────────────────────────────

@login_required
def article_toggle_status(request, pk):
    """Habilitar/Deshabilitar artículo (AJAX)"""
    if request.method == 'POST':
        try:
            article = Article.objects.get(pk=pk)
            if article.status == 'available':
                article.status = 'retired'
                message = f'{article.name} deshabilitado'
            else:
                article.status = 'available'
                message = f'{article.name} habilitado'
            article.save()
            return JsonResponse({
                'success': True,
                'message': message,
                'new_status': article.status,
                'status_display': article.get_status_display(),
            })
        except Article.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Artículo no encontrado'}, status=404)

    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)