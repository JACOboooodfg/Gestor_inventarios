from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, Sum, Count, F, FloatField, ExpressionWrapper
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.contrib.auth.models import User
from .models import Category, Location, Article, Movement, Loan, Alert
from .forms import (LoginForm, CategoryForm, LocationForm, ArticleForm, 
                   MovementForm, LoanForm, ImportExcelForm, ArticleSearchForm)
from .utils import (export_articles_to_excel, import_articles_from_excel,
                   export_movements_to_excel, export_loans_to_excel)


# ==================== AUTENTICACIÓN ====================

def register_view(request):
    """Vista de registro de nuevos usuarios"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        
        # Validaciones
        if not all([username, email, password1, password2]):
            messages.error(request, 'Todos los campos son requeridos')
            return render(request, 'inventory/register.html')
        
        if password1 != password2:
            messages.error(request, 'Las contraseñas no coinciden')
            return render(request, 'inventory/register.html')
        
        if len(password1) < 6:
            messages.error(request, 'La contraseña debe tener al menos 6 caracteres')
            return render(request, 'inventory/register.html')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'El nombre de usuario ya existe')
            return render(request, 'inventory/register.html')
        
        if User.objects.filter(email=email).exists():
            messages.error(request, 'El email ya está registrado')
            return render(request, 'inventory/register.html')
        
        # Crear usuario
        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
                last_name=last_name
            )
            # Login con backend especificado
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            messages.success(request, f'¡Bienvenido {username}! Tu cuenta ha sido creada exitosamente.')
            return redirect('select_inventory')
        except Exception as e:
            messages.error(request, f'Error al crear la cuenta: {str(e)}')
            return render(request, 'inventory/register.html')
    
    return render(request, 'inventory/register.html')


def login_view(request):
    """Vista de login"""
    if request.user.is_authenticated:
        return redirect('select_inventory')  # Redirigir a pantalla de selección
    
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Bienvenido, {user.username}!')
                return redirect('select_inventory')  # Redirigir a pantalla de selección
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    else:
        form = LoginForm()
    
    # Verificar si Google OAuth está configurado
    google_login_enabled = False
    try:
        from django.conf import settings
        google_login_enabled = 'social_django' in settings.INSTALLED_APPS
    except:
        pass
    
    return render(request, 'inventory/login.html', {
        'form': form,
        'google_login_enabled': google_login_enabled
    })


@login_required
def logout_view(request):
    """Vista de logout"""
    logout(request)
    messages.info(request, 'Has cerrado sesión exitosamente')
    return redirect('login')


# ==================== DASHBOARD ====================

@login_required
def dashboard(request):
    """Dashboard principal con estadísticas"""
    # Estadísticas generales
    total_articles = Article.objects.count()
    total_categories = Category.objects.count()
    total_locations = Location.objects.count()
    low_stock_count = Article.objects.filter(quantity__lte=F('min_quantity')).count()
    
    # Artículos con stock bajo
    low_stock_articles = Article.objects.filter(
        quantity__lte=F('min_quantity')
    ).select_related('category', 'location')[:10]
    
    # Préstamos activos y vencidos
    active_loans = Loan.objects.filter(status='active').count()
    overdue_loans = Loan.objects.filter(
        status='active',
        due_date__lt=timezone.now()
    ).count()
    
    # Movimientos recientes
    recent_movements = Movement.objects.select_related(
        'article', 'user', 'article__category'
    ).order_by('-created_at')[:10]
    
    # Alertas no leídas
    unread_alerts = Alert.objects.filter(is_read=False).order_by('-created_at')[:5]
    
    # Estadísticas por categoría
    category_stats = Category.objects.all().order_by('-created_at')[:5]


    context = {
        'total_articles': total_articles,
        'total_categories': total_categories,
        'total_locations': total_locations,
        'low_stock_count': low_stock_count,
        'low_stock_articles': low_stock_articles,
        'active_loans': active_loans,
        'overdue_loans': overdue_loans,
        'recent_movements': recent_movements,
        'unread_alerts': unread_alerts,
        'category_stats': category_stats,
    }
    
    return render(request, 'inventory/dashboard.html', context)


# ==================== CATEGORÍAS ====================

@login_required
def category_list(request):
    """Lista de categorías"""
    categories = Category.objects.all().order_by('name')
    
    return render(request, 'inventory/category_list.html', {
        'categories': categories
    })


@login_required
def category_create(request):
    """Crear nueva categoría"""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Categoría "{category.name}" creada exitosamente')
            return redirect('category_list')
    else:
        form = CategoryForm()
    
    return render(request, 'inventory/category_form.html', {
        'form': form,
        'title': 'Nueva Categoría'
    })


@login_required
def category_update(request, pk):
    """Actualizar categoría"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, f'Categoría "{category.name}" actualizada')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    
    return render(request, 'inventory/category_form.html', {
        'form': form,
        'title': f'Editar: {category.name}'
    })


@login_required
def category_delete(request, pk):
    """Eliminar categoría"""
    category = get_object_or_404(Category, pk=pk)
    
    if category.articles.exists():
        messages.error(request, 'No se puede eliminar una categoría con artículos asignados')
        return redirect('category_list')
    
    if request.method == 'POST':
        name = category.name
        category.delete()
        messages.success(request, f'Categoría "{name}" eliminada')
        return redirect('category_list')
    
    return render(request, 'inventory/category_confirm_delete.html', {
        'category': category
    })


@login_required
def category_inventory(request, pk):
    """Inventario específico de una categoría"""
    category = get_object_or_404(Category, pk=pk)
    
    # Obtener artículos de esta categoría
    articles = Article.objects.filter(category=category).select_related(
        'location', 'created_by'
    )
    
    # Búsqueda dentro de la categoría
    search_query = request.GET.get('q', '')
    if search_query:
        articles = articles.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filtro por estado
    status_filter = request.GET.get('status', '')
    if status_filter:
        articles = articles.filter(status=status_filter)
    
    # Filtro por ubicación
    location_filter = request.GET.get('location', '')
    if location_filter:
        articles = articles.filter(location_id=location_filter)
    
    # Ordenamiento
    sort_by = request.GET.get('sort', '-created_at')
    articles = articles.order_by(sort_by)
    
    # Estadísticas de la categoría
    stats = articles.aggregate(
        total_value=Sum(ExpressionWrapper(F('quantity') * F('price'), output_field=FloatField())),
        total_quantity=Sum('quantity')
    )
    total_value = round(float(stats['total_value'] or 0), 2)
    total_quantity = int(stats['total_quantity'] or 0)
    total_items = articles.count()
    low_stock_count = articles.filter(quantity__lte=F('min_quantity')).count()

    
    context = {
        'category': category,
        'articles': articles,
        'total_items': total_items,
        'total_quantity': total_quantity,
        'low_stock_count': low_stock_count,
        'total_value': total_value,
        'search_query': search_query,
        'locations': Location.objects.all(),
    }
    
    return render(request, 'inventory/category_inventory.html', context)


@login_required
def category_export(request, pk):
    """Exportar artículos de una categoría específica a Excel"""
    category = get_object_or_404(Category, pk=pk)
    articles = Article.objects.filter(category=category).select_related('location')
    
    # Aplicar filtros si existen
    search_query = request.GET.get('q', '')
    if search_query:
        articles = articles.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    status_filter = request.GET.get('status', '')
    if status_filter:
        articles = articles.filter(status=status_filter)
    
    location_filter = request.GET.get('location', '')
    if location_filter:
        articles = articles.filter(location_id=location_filter)
    
    return export_articles_to_excel(articles)


@login_required
def category_import(request, pk):
    """Importar artículos a una categoría específica"""
    category = get_object_or_404(Category, pk=pk)
    
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                from openpyxl import load_workbook
                
                wb = load_workbook(excel_file)
                ws = wb.active
                
                success_count = 0
                error_count = 0
                errors = []
                
                # Leer encabezados
                headers = [str(cell.value).lower().strip() for cell in ws[1] if cell.value]
                
                # Procesar cada fila
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not any(row):  # Saltar filas vacías
                            continue
                        
                        # Construir diccionario con los datos
                        data = {}
                        for idx, header in enumerate(headers):
                            if idx < len(row):
                                data[header] = row[idx]
                        
                        # Validar nombre
                        if not data.get('nombre'):
                            errors.append(f"Fila {row_num}: Nombre es requerido")
                            error_count += 1
                            continue
                        
                        # Procesar ubicación
                        location = None
                        if data.get('ubicacion'):
                            loc_name = str(data['ubicacion']).strip()
                            try:
                                location = Location.objects.get(name__iexact=loc_name)
                            except Location.DoesNotExist:
                                pass
                        
                        # Crear artículo - ASIGNAR AUTOMÁTICAMENTE LA CATEGORÍA
                        article_data = {
                            'name': str(data['nombre']).strip(),
                            'category': category,  # ← IMPORTANTE: Asignar la categoría actual
                            'location': location,
                            'quantity': int(data.get('cantidad', 0)),
                            'unit': str(data.get('unidad', 'unidad')).strip(),
                            'min_quantity': int(data.get('cantidad_minima', 5)),
                            'status': data.get('estado', 'available'),
                            'description': str(data.get('descripcion', '')).strip(),
                            'notes': str(data.get('notas', '')).strip(),
                            'created_by': request.user,
                        }
                        
                        # Código (si existe)
                        if data.get('codigo'):
                            article_data['code'] = str(data['codigo']).strip()
                        
                        # Precio
                        if data.get('precio'):
                            try:
                                article_data['price'] = float(data['precio'])
                            except (ValueError, TypeError):
                                pass
                        
                        # Código de barras
                        if data.get('codigo_barras'):
                            article_data['barcode'] = str(data['codigo_barras']).strip()
                        
                        # Crear artículo
                        Article.objects.create(**article_data)
                        success_count += 1
                        
                    except Exception as e:
                        errors.append(f"Fila {row_num}: {str(e)}")
                        error_count += 1
                
                # Mensajes de resultado
                if success_count > 0:
                    messages.success(request, f'✓ {success_count} artículos importados exitosamente a {category.name}')
                
                if error_count > 0:
                    messages.warning(request, f'⚠ {error_count} filas con errores')
                    for error in errors[:10]:  # Mostrar solo los primeros 10 errores
                        messages.error(request, error)
                
                if success_count > 0:
                    return redirect('category_inventory', pk=pk)
                
            except Exception as e:
                messages.error(request, f'Error al procesar archivo: {str(e)}')
    else:
        form = ImportExcelForm()
    
    return render(request, 'inventory/category_import.html', {
        'form': form,
        'category': category
    })


# ==================== UBICACIONES ====================

@login_required
def location_list(request):
    """Lista de ubicaciones"""
    locations = Location.objects.all().order_by('building', 'floor', 'room')
    
    return render(request, 'inventory/location_list.html', {
        'locations': locations
    })


@login_required
def location_create(request):
    """Crear nueva ubicación"""
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            location = form.save()
            messages.success(request, f'Ubicación "{location.name}" creada exitosamente')
            return redirect('location_list')
    else:
        form = LocationForm()
    
    return render(request, 'inventory/location_form.html', {
        'form': form,
        'title': 'Nueva Ubicación'
    })


@login_required
def location_update(request, pk):
    """Actualizar ubicación"""
    location = get_object_or_404(Location, pk=pk)
    
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            messages.success(request, f'Ubicación "{location.name}" actualizada')
            return redirect('location_list')
    else:
        form = LocationForm(instance=location)
    
    return render(request, 'inventory/location_form.html', {
        'form': form,
        'title': f'Editar: {location.name}'
    })


@login_required
def location_delete(request, pk):
    """Eliminar ubicación"""
    location = get_object_or_404(Location, pk=pk)
    
    if request.method == 'POST':
        name = location.name
        location.delete()
        messages.success(request, f'Ubicación "{name}" eliminada')
        return redirect('location_list')
    
    return render(request, 'inventory/location_confirm_delete.html', {
        'location': location
    })


# ==================== ARTÍCULOS ====================

@login_required
def article_list(request):
    """Lista de artículos con búsqueda y filtros"""
    from django.core.paginator import Paginator
    
    articles = Article.objects.select_related(
        'category', 'location', 'created_by'
    ).only(
        'id', 'code', 'name', 'description', 'quantity', 'min_quantity',
        'unit', 'status', 'image', 'purchase_date', 'supplier',
        'category__id', 'category__name', 'category__color',
        'location__id', 'location__name',
        'created_by__id','barcode'
    )
    
    # Búsqueda con nuevo formulario
    search_form = ArticleSearchForm(request.GET)
    
    if search_form.is_valid():
        query = search_form.cleaned_data.get('query')
        category = search_form.cleaned_data.get('category')
        status = search_form.cleaned_data.get('status')
        location = search_form.cleaned_data.get('location')
        fecha_desde = search_form.cleaned_data.get('fecha_desde')
        fecha_hasta = search_form.cleaned_data.get('fecha_hasta')
        proveedor = search_form.cleaned_data.get('proveedor')
        
        if query:
            articles = articles.filter(
                Q(name__icontains=query) |
                Q(code__icontains=query) |
                Q(description__icontains=query) |
                Q(barcode__icontains=query)
            )
        
        if category:
            articles = articles.filter(category=category)
        
        if status:
            articles = articles.filter(status=status)
        
        if location:
            articles = articles.filter(location=location)
        
        if fecha_desde:
            articles = articles.filter(purchase_date__gte=fecha_desde)
        
        if fecha_hasta:
            articles = articles.filter(purchase_date__lte=fecha_hasta)
        
        if proveedor:
            articles = articles.filter(supplier__icontains=proveedor)
    
    # Ordenamiento
    sort_by = request.GET.get('sort', '-created_at')
    articles = articles.order_by(sort_by)
    
    # Paginación
    paginator = Paginator(articles, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'articles': page_obj,
        'page_obj': page_obj,
        'search_form': search_form,
        'total_count': paginator.count,
    }
    
    return render(request, 'inventory/article_list.html', context)

@login_required
def article_detail(request, pk):
    """Detalle de artículo"""
    article = get_object_or_404(
        Article.objects.select_related('category', 'location', 'created_by'),
        pk=pk
    )
    
    # Movimientos del artículo
    movements = article.movements.select_related('user').order_by('-created_at')[:20]
    
    # Préstamos del artículo
    loans = article.loans.select_related('approved_by', 'returned_by').order_by('-loan_date')[:10]
    
    context = {
        'article': article,
        'movements': movements,
        'loans': loans,
    }
    
    return render(request, 'inventory/article_detail.html', context)


@login_required
def article_create(request):
    """Crear nuevo artículo"""
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES)
        if form.is_valid():
            article = form.save(commit=False)
            article.created_by = request.user
            article.save()
            
            # Registrar movimiento inicial
            if article.quantity > 0:
                Movement.objects.create(
                    article=article,
                    movement_type='entry',
                    quantity=article.quantity,
                    previous_quantity=0,
                    new_quantity=article.quantity,
                    reason='Registro inicial de inventario',
                    user=request.user
                )
            
            messages.success(request, f'Artículo "{article.name}" creado exitosamente')
            
            # Redirigir al inventario de la categoría si vino desde ahí
            from_category = request.POST.get('from_category')
            if from_category:
                return redirect('category_inventory', pk=from_category)
            
            return redirect('article_detail', pk=article.pk)
    else:
        form = ArticleForm()
        
        # Pre-seleccionar categoría si viene del parámetro
        category_id = request.GET.get('category')
        if category_id:
            form.initial['category'] = category_id
    
    return render(request, 'inventory/article_form.html', {
        'form': form,
        'title': 'Nuevo Artículo',
        'from_category': request.GET.get('category', '')
    })


@login_required
def article_update(request, pk):
    """Actualizar artículo"""
    article = get_object_or_404(Article, pk=pk)
    old_quantity = article.quantity
    
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES, instance=article)
        if form.is_valid():
            article = form.save()
            
            # Si cambió la cantidad, registrar movimiento
            if article.quantity != old_quantity:
                movement_type = 'entry' if article.quantity > old_quantity else 'exit'
                quantity_diff = abs(article.quantity - old_quantity)
                
                Movement.objects.create(
                    article=article,
                    movement_type='adjustment',
                    quantity=quantity_diff,
                    previous_quantity=old_quantity,
                    new_quantity=article.quantity,
                    reason=f'Ajuste manual desde edición de artículo',
                    user=request.user
                )
            
            messages.success(request, f'Artículo "{article.name}" actualizado')
            return redirect('article_detail', pk=article.pk)
    else:
        form = ArticleForm(instance=article)
    
    return render(request, 'inventory/article_form.html', {
        'form': form,
        'title': f'Editar: {article.name}'
    })


@login_required
def article_delete(request, pk):
    """Eliminar artículo"""
    article = get_object_or_404(Article, pk=pk)
    
    if request.method == 'POST':
        name = article.name
        article.delete()
        messages.success(request, f'Artículo "{name}" eliminado')
        return redirect('article_list')
    
    return render(request, 'inventory/article_confirm_delete.html', {
        'article': article
    })


# ==================== MOVIMIENTOS ====================

@login_required
def movement_list(request):
    """Lista de movimientos"""
    movements = Movement.objects.select_related(
        'article', 'article__category', 'user', 'from_location', 'to_location'
    ).order_by('-created_at')
    
    # Filtros
    movement_type = request.GET.get('type')
    if movement_type:
        movements = movements.filter(movement_type=movement_type)
    
    date_from = request.GET.get('date_from')
    if date_from:
        movements = movements.filter(created_at__gte=date_from)
    
    date_to = request.GET.get('date_to')
    if date_to:
        movements = movements.filter(created_at__lte=date_to)
    
    context = {
        'movements': movements,
        'total_count': movements.count(),
    }
    
    return render(request, 'inventory/movement_list.html', context)


@login_required
def movement_create(request):
    """Registrar nuevo movimiento"""
    if request.method == 'POST':
        form = MovementForm(request.POST)
        if form.is_valid():
            movement = form.save(commit=False)
            article = movement.article
            
            # Guardar cantidad anterior
            movement.previous_quantity = article.quantity
            
            # Calcular nueva cantidad según tipo de movimiento
            if movement.movement_type in ['entry', 'return']:
                article.quantity += movement.quantity
            elif movement.movement_type in ['exit', 'loan']:
                if article.quantity < movement.quantity:
                    messages.error(request, 'No hay suficiente cantidad disponible')
                    return render(request, 'inventory/movement_form.html', {
                        'form': form,
                        'title': 'Nuevo Movimiento'
                    })
                article.quantity -= movement.quantity
            elif movement.movement_type == 'adjustment':
                # El ajuste se hace directamente
                pass
            
            movement.new_quantity = article.quantity
            movement.user = request.user
            
            # Guardar
            article.save()
            movement.save()
            
            messages.success(request, 'Movimiento registrado exitosamente')
            return redirect('movement_list')
    else:
        form = MovementForm()
    
    return render(request, 'inventory/movement_form.html', {
        'form': form,
        'title': 'Nuevo Movimiento'
    })


# ==================== PRÉSTAMOS ====================

@login_required
def loan_list(request):
    """Lista de préstamos"""
    loans = Loan.objects.select_related(
        'article', 'article__category', 'approved_by', 'returned_by'
    ).order_by('-loan_date')
    
    # Filtros
    status = request.GET.get('status')
    if status:
        loans = loans.filter(status=status)
    
    # Actualizar estado de préstamos vencidos
    for loan in loans:
        if loan.is_overdue and loan.status == 'active':
            loan.status = 'overdue'
            loan.save()
    
    context = {
        'loans': loans,
        'total_count': loans.count(),
        'active_count': loans.filter(status='active').count(),
        'overdue_count': loans.filter(status='overdue').count(),
    }
    
    return render(request, 'inventory/loan_list.html', context)


@login_required
def loan_create(request):
    """Crear nuevo préstamo"""
    if request.method == 'POST':
        form = LoanForm(request.POST)
        if form.is_valid():
            loan = form.save(commit=False)
            article = loan.article
            
            # Verificar disponibilidad
            if article.quantity < loan.quantity:
                messages.error(request, f'No hay suficiente cantidad disponible. Disponible: {article.quantity}')
                return render(request, 'inventory/loan_form.html', {
                    'form': form,
                    'title': 'Nuevo Préstamo'
                })
            
            # Reducir cantidad del artículo
            article.quantity -= loan.quantity
            article.save()
            
            # Guardar préstamo
            loan.approved_by = request.user
            loan.save()
            
            # Registrar movimiento
            Movement.objects.create(
                article=article,
                movement_type='loan',
                quantity=loan.quantity,
                previous_quantity=article.quantity + loan.quantity,
                new_quantity=article.quantity,
                reason=f'Préstamo a {loan.borrower_name}',
                reference=f'LOAN-{loan.id}',
                user=request.user
            )
            
            messages.success(request, 'Préstamo registrado exitosamente')
            return redirect('loan_list')
    else:
        form = LoanForm()
    
    return render(request, 'inventory/loan_form.html', {
        'form': form,
        'title': 'Nuevo Préstamo'
    })


@login_required
def loan_return(request, pk):
    """Devolver préstamo"""
    loan = get_object_or_404(Loan, pk=pk)
    
    if loan.status != 'active' and loan.status != 'overdue':
        messages.error(request, 'Este préstamo ya fue devuelto o cancelado')
        return redirect('loan_list')
    
    if request.method == 'POST':
        # Actualizar préstamo
        loan.return_date = timezone.now()
        loan.status = 'returned'
        loan.returned_by = request.user
        loan.save()
        
        # Devolver cantidad al artículo
        article = loan.article
        article.quantity += loan.quantity
        article.save()
        
        # Registrar movimiento
        Movement.objects.create(
            article=article,
            movement_type='return',
            quantity=loan.quantity,
            previous_quantity=article.quantity - loan.quantity,
            new_quantity=article.quantity,
            reason=f'Devolución de préstamo de {loan.borrower_name}',
            reference=f'LOAN-{loan.id}',
            user=request.user
        )
        
        messages.success(request, 'Préstamo devuelto exitosamente')
        return redirect('loan_list')
    
    return render(request, 'inventory/loan_return.html', {
        'loan': loan
    })


# ==================== IMPORTAR/EXPORTAR ====================

@login_required
def import_articles(request):
    """Importar artículos desde Excel (VERSIÓN LEGACY - mantener para compatibilidad)"""
    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                results = import_articles_from_excel(excel_file)
                
                # Mostrar resultados
                for success_msg in results['success']:
                    messages.success(request, success_msg)
                
                for warning_msg in results['warnings']:
                    messages.warning(request, warning_msg)
                
                for error_msg in results['errors']:
                    messages.error(request, error_msg)
                
                if results['success']:
                    return redirect('article_list')
                
            except Exception as e:
                messages.error(request, f'Error al procesar archivo: {str(e)}')
    else:
        form = ImportExcelForm()
    
    return render(request, 'inventory/import_articles.html', {
        'form': form
    })


@login_required
def export_articles(request):
    """Exportar artículos a Excel"""
    articles = Article.objects.select_related('category', 'location').all()
    
    # Aplicar los mismos filtros que en la lista
    search_form = SearchForm(request.GET)
    if search_form.is_valid():
        query = search_form.cleaned_data.get('query')
        category = search_form.cleaned_data.get('category')
        status = search_form.cleaned_data.get('status')
        location = search_form.cleaned_data.get('location')
        
        if query:
            articles = articles.filter(
                Q(name__icontains=query) |
                Q(code__icontains=query) |
                Q(description__icontains=query)
            )
        if category:
            articles = articles.filter(category=category)
        if status:
            articles = articles.filter(status=status)
        if location:
            articles = articles.filter(location=location)
    
    return export_articles_to_excel(articles)


@login_required
def export_movements(request):
    """Exportar movimientos a Excel"""
    movements = Movement.objects.select_related(
        'article', 'user', 'from_location', 'to_location'
    ).all()
    
    # Aplicar filtros si existen
    movement_type = request.GET.get('type')
    if movement_type:
        movements = movements.filter(movement_type=movement_type)
    
    return export_movements_to_excel(movements)


@login_required
def export_loans(request):
    """Exportar préstamos a Excel"""
    loans = Loan.objects.select_related(
        'article', 'approved_by', 'returned_by'
    ).all()
    
    status = request.GET.get('status')
    if status:
        loans = loans.filter(status=status)
    
    return export_loans_to_excel(loans)


# ==================== ALERTAS ====================

@login_required
def alert_list(request):
    """Lista de alertas"""
    alerts = Alert.objects.select_related(
        'article', 'loan', 'read_by'
    ).order_by('-created_at')
    
    # Filtrar por leídas/no leídas
    filter_read = request.GET.get('read')
    if filter_read == 'unread':
        alerts = alerts.filter(is_read=False)
    elif filter_read == 'read':
        alerts = alerts.filter(is_read=True)
    
    context = {
        'alerts': alerts,
        'unread_count': Alert.objects.filter(is_read=False).count(),
    }
    
    return render(request, 'inventory/alert_list.html', context)


@login_required
def alert_mark_read(request, pk):
    """Marcar alerta como leída"""
    alert = get_object_or_404(Alert, pk=pk)
    
    if not alert.is_read:
        alert.is_read = True
        alert.read_by = request.user
        alert.read_at = timezone.now()
        alert.save()
    
    return redirect('alert_list')


@login_required
def alert_mark_all_read(request):
    """Marcar todas las alertas como leídas"""
    Alert.objects.filter(is_read=False).update(
        is_read=True,
        read_by=request.user,
        read_at=timezone.now()
    )
    
    messages.success(request, 'Todas las alertas marcadas como leídas')
    return redirect('alert_list')


# ==================== REPORTES ====================

@login_required
def reports(request):
    """Vista de reportes y estadísticas"""
    # Estadísticas por categoría
    category_report = Category.objects.all().order_by('name')
    
    # Movimientos por tipo (último mes)
    last_month = timezone.now() - timedelta(days=30)
    movement_stats = Movement.objects.filter(
        created_at__gte=last_month
    ).values('movement_type').annotate(
        count=Count('id'),
        total_quantity=Sum('quantity')
    )
    
    # Artículos más prestados
    top_loaned = Article.objects.annotate(
        loan_count=Count('loans')
    ).filter(loan_count__gt=0).order_by('-loan_count')[:10]
    
    # Préstamos por estado
    loan_stats = Loan.objects.values('status').annotate(
        count=Count('id')
    )
    
    context = {
        'category_report': category_report,
        'movement_stats': movement_stats,
        'top_loaned': top_loaned,
        'loan_stats': loan_stats,
    }
    
    return render(request, 'inventory/reports.html', context)


@login_required
def delete_all_articles_confirm(request):
    """Confirmar antes de borrar todos los artículos"""
    total_articles = Article.objects.count()
    
    context = {
        'total_articles': total_articles,
    }
    
    return render(request, 'inventory/delete_all_confirm.html', context)


@login_required
def delete_all_articles_execute(request):
    """Ejecutar el borrado de todos los artículos"""
    if request.method == 'POST':
        # Doble verificación
        confirm_text = request.POST.get('confirm_text', '')
        
        if confirm_text.upper() == 'BORRAR TODO':
            try:
                total = Article.objects.count()
                Article.objects.all().delete()
                messages.success(request, f'✓ {total} artículos eliminados exitosamente. Base de datos limpia.')
                return redirect('dashboard')
            except Exception as e:
                messages.error(request, f'Error al eliminar: {str(e)}')
                return redirect('delete_all_articles_confirm')
        else:
            messages.error(request, 'Texto de confirmación incorrecto. Debes escribir exactamente: BORRAR TODO')
            return redirect('delete_all_articles_confirm')
    
    return redirect('delete_all_articles_confirm')

# ==================== NUEVAS VISTAS - AGREGAR AL FINAL DE views.py ====================

@login_required
def select_inventory(request):
    """Pantalla de selección entre Inventario General e Inventario de Aseo"""
    return render(request, 'inventory/select_inventory.html')


@login_required
def category_list_with_locations(request):
    """Lista de categorías con sus ubicaciones anidadas"""
    from django.db.models import Count, Q
    
    categories = Category.objects.annotate(
        total_items=Count('articles'),
        total_quantity=Sum('articles__quantity')
    ).prefetch_related('articles__location').all()
    
    # Para cada categoría, obtener sus ubicaciones con conteo
    for category in categories:
        # Obtener ubicaciones únicas que tienen artículos de esta categoría
        locations_with_count = Location.objects.filter(
            articles__category=category
        ).annotate(
            article_count=Count('articles', filter=Q(articles__category=category))
        ).distinct()
        
        category.locations = locations_with_count
    
    context = {
        'categories': categories,
        'total_categories': categories.count(),
    }
    
    return render(request, 'inventory/category_list_with_locations.html', context)


@login_required
def category_location_articles(request, category_pk, location_pk):
    """Artículos de una categoría en una ubicación específica"""
    category = get_object_or_404(Category, pk=category_pk)
    location = get_object_or_404(Location, pk=location_pk)
    
    articles = Article.objects.filter(
        category=category,
        location=location
    ).select_related('category', 'location', 'created_by')
    
    context = {
        'articles': articles,
        'category': category,
        'location': location,
        'total_count': articles.count(),
    }
    
    return render(request, 'inventory/article_list.html', context)


@login_required
def aseo_dashboard(request):
    """Dashboard simplificado solo para artículos de aseo"""
    # Obtener categoría de Aseo y Limpieza
    try:
        aseo_category = Category.objects.get(name__icontains='aseo')
    except Category.DoesNotExist:
        aseo_category = None
    
    # Filtrar solo artículos de aseo
    productos = Article.objects.filter(
        category=aseo_category
    ).select_related('category', 'location') if aseo_category else Article.objects.none()
    
    # Búsqueda simple
    query = request.GET.get('q', '')
    if query:
        productos = productos.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(description__icontains=query)
        )
    
    # Estadísticas
    total_productos = productos.count()
    disponibles = productos.filter(status='available').count()
    stock_bajo = productos.filter(quantity__lte=F('min_quantity')).count()
    ubicaciones = Location.objects.filter(
        articles__category=aseo_category
    ).distinct().count() if aseo_category else 0
    
    context = {
        'productos': productos,
        'query': query,
        'total_productos': total_productos,
        'disponibles': disponibles,
        'stock_bajo': stock_bajo,
        'ubicaciones': ubicaciones,
        'locations': Location.objects.all(),        
        
    }
    
    return render(request, 'inventory/aseo_dashboard.html', context)

# ==================== NUEVA VISTA PARA AJUSTAR CANTIDAD ====================
# Agregar a inventory/views.py

@login_required
def aseo_ajustar_cantidad(request, pk):
    """Ajustar cantidad de producto de aseo con botones +/-"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            cambio = data.get('cambio', 0)  # +1 o -1
            
            article = get_object_or_404(Article, pk=pk)
            
            # Calcular nueva cantidad
            nueva_cantidad = article.quantity + cambio
            
            # Validar que no sea negativa
            if nueva_cantidad < 0:
                return JsonResponse({
                    'success': False,
                    'message': 'La cantidad no puede ser negativa'
                })
            
            # Actualizar cantidad
            article.quantity = nueva_cantidad
            article.save()
            
            # Determinar mensaje
            if cambio > 0:
                mensaje = f'+{cambio} {article.unit} agregado'
            else:
                mensaje = f'{abs(cambio)} {article.unit} removido'
            
            return JsonResponse({
                'success': True,
                'message': mensaje,
                'nueva_cantidad': nueva_cantidad,
                'is_low_stock': article.is_low_stock
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

@login_required
def aseo_crear_producto(request):
    """Crear producto directamente desde el dashboard de aseo"""
    if request.method == 'POST':
        try:
            # Obtener o buscar categoría de aseo
            aseo_category = Category.objects.filter(name__icontains='aseo').first()
            
            name = request.POST.get('name', '').strip()
            if not name:
                messages.error(request, 'El nombre es requerido')
                return redirect('aseo_dashboard')
            
            quantity = int(request.POST.get('quantity', 0))
            unit = request.POST.get('unit', 'unidad')
            min_quantity = int(request.POST.get('min_quantity', 5))
            location_id = request.POST.get('location')
            
            location = None
            if location_id:
                location = Location.objects.filter(pk=location_id).first()
            
            article = Article.objects.create(
                name=name,
                category=aseo_category,
                quantity=quantity,
                unit=unit,
                min_quantity=min_quantity,
                location=location,
                created_by=request.user,
                status='available'
            )
            
            # Movimiento inicial
            if quantity > 0:
                Movement.objects.create(
                    article=article,
                    movement_type='entry',
                    quantity=quantity,
                    previous_quantity=0,
                    new_quantity=quantity,
                    reason='Registro inicial desde inventario de aseo',
                    user=request.user
                )
            
            messages.success(request, f'Producto "{name}" creado exitosamente')
        
        except Exception as e:
            messages.error(request, f'Error al crear producto: {str(e)}')
    
    return redirect('aseo_dashboard')

@login_required
def aseo_eliminar_producto(request, pk):
    """Eliminar producto desde el dashboard de aseo"""
    article = get_object_or_404(Article, pk=pk)
    
    if request.method == 'POST':
        name = article.name
        article.delete()
        messages.success(request, f'Producto "{name}" eliminado exitosamente')
        return redirect('aseo_dashboard')
    
    return redirect('aseo_dashboard')



# ==================== NOTA IMPORTANTE ====================
# Las funciones de importación especializada para el colegio están en import_views.py:
# - import_preview() - Importación con 15 hojas, lectura desde fila 8, detección de X
# - import_colegio_excel() - Lógica especializada
# - article_toggle_status() - Toggle de estado
# 
# Ver inventory/import_views.py para el código completo