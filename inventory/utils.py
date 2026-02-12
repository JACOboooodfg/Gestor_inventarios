from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from .models import Article, Category, Location
from datetime import datetime


def export_articles_to_excel(articles):
    """Exportar artículos a Excel con formato profesional"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'Código', 'Nombre', 'Categoría', 'Ubicación', 'Cantidad', 
        'Unidad', 'Cantidad Mínima', 'Estado', 'Precio', 'Valor Total', 
        'Código de Barras', 'Descripción', 'Notas', 'Fecha de Creación'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos
    for row, article in enumerate(articles, start=2):
        data = [
            article.code,
            article.name,
            article.category.name,
            str(article.location) if article.location else '',
            article.quantity,
            article.unit,
            article.min_quantity,
            article.get_status_display(),
            float(article.price) if article.price else 0,
            float(article.total_value) if article.total_value else 0,
            article.barcode,
            article.description,
            article.notes,
            article.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [5, 7, 9, 10]:  # Columnas numéricas
                cell.alignment = Alignment(horizontal="right")
    
    # Ajustar anchos de columna
    column_widths = [15, 30, 20, 25, 12, 12, 15, 15, 12, 15, 20, 40, 40, 20]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=inventario_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


def import_articles_from_excel(excel_file):
    """Importar artículos desde Excel
    
    Columnas esperadas:
    - codigo (opcional, se auto-genera si está vacío)
    - nombre (requerido)
    - categoria (requerido, debe existir)
    - ubicacion (opcional, debe existir si se proporciona)
    - cantidad (requerido, número)
    - unidad (opcional, default: 'unidad')
    - cantidad_minima (opcional, default: 5)
    - estado (opcional, default: 'available')
    - precio (opcional)
    - codigo_barras (opcional)
    - descripcion (opcional)
    - notas (opcional)
    """
    wb = load_workbook(excel_file)
    ws = wb.active
    
    results = {
        'success': [],
        'errors': [],
        'warnings': []
    }
    
    # Leer encabezados (primera fila)
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).lower().strip().replace(' ', '_'))
    
    # Mapeo de nombres alternativos de columnas
    column_map = {
        'codigo': ['codigo', 'code', 'cod'],
        'nombre': ['nombre', 'name', 'articulo', 'item'],
        'categoria': ['categoria', 'category', 'cat'],
        'ubicacion': ['ubicacion', 'location', 'lugar'],
        'cantidad': ['cantidad', 'quantity', 'stock', 'qty'],
        'unidad': ['unidad', 'unit', 'medida'],
        'cantidad_minima': ['cantidad_minima', 'min_quantity', 'minimo'],
        'estado': ['estado', 'status'],
        'precio': ['precio', 'price', 'costo'],
        'codigo_barras': ['codigo_barras', 'barcode', 'ean'],
        'descripcion': ['descripcion', 'description', 'desc'],
        'notas': ['notas', 'notes', 'observaciones']
    }
    
    # Crear índice de columnas
    col_index = {}
    for standard_name, alternatives in column_map.items():
        for idx, header in enumerate(headers):
            if header in alternatives:
                col_index[standard_name] = idx
                break
    
    # Validar columnas requeridas
    required_cols = ['nombre', 'categoria', 'cantidad']
    missing_cols = [col for col in required_cols if col not in col_index]
    if missing_cols:
        results['errors'].append(f"Columnas requeridas faltantes: {', '.join(missing_cols)}")
        return results
    
    # Procesar filas (empezando desde la segunda)
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Extraer datos
            data = {}
            for field, idx in col_index.items():
                data[field] = row[idx] if idx < len(row) else None
            
            # Validar datos requeridos
            if not data.get('nombre'):
                results['errors'].append(f"Fila {row_num}: Nombre es requerido")
                continue
            
            if not data.get('categoria'):
                results['errors'].append(f"Fila {row_num}: Categoría es requerida")
                continue
            
            # Buscar o crear categoría
            category_name = str(data['categoria']).strip()
            category, created = Category.objects.get_or_create(
                name__iexact=category_name,
                defaults={'name': category_name}
            )
            if created:
                results['warnings'].append(f"Fila {row_num}: Categoría '{category_name}' creada automáticamente")
            
            # Buscar ubicación si se proporciona
            location = None
            if data.get('ubicacion'):
                location_name = str(data['ubicacion']).strip()
                try:
                    location = Location.objects.get(name__iexact=location_name)
                except Location.DoesNotExist:
                    results['warnings'].append(f"Fila {row_num}: Ubicación '{location_name}' no encontrada, se omitirá")
            
            # Crear artículo
            article_data = {
                'name': str(data['nombre']).strip(),
                'category': category,
                'location': location,
                'quantity': int(data.get('cantidad', 0)),
                'unit': str(data.get('unidad', 'unidad')).strip(),
                'min_quantity': int(data.get('cantidad_minima', 5)),
                'status': data.get('estado', 'available'),
                'description': str(data.get('descripcion', '')).strip() if data.get('descripcion') else '',
                'notes': str(data.get('notas', '')).strip() if data.get('notas') else '',
            }
            
            # Campos opcionales
            if data.get('codigo'):
                article_data['code'] = str(data['codigo']).strip()
            
            if data.get('precio'):
                try:
                    article_data['price'] = float(data['precio'])
                except (ValueError, TypeError):
                    pass
            
            if data.get('codigo_barras'):
                article_data['barcode'] = str(data['codigo_barras']).strip()
            
            # Crear artículo
            article = Article.objects.create(**article_data)
            results['success'].append(f"Fila {row_num}: Artículo '{article.name}' creado exitosamente (Código: {article.code})")
            
        except Exception as e:
            results['errors'].append(f"Fila {row_num}: Error - {str(e)}")
    
    return results


def export_movements_to_excel(movements):
    """Exportar movimientos a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'Fecha', 'Tipo', 'Artículo', 'Cantidad', 'Cantidad Anterior', 
        'Cantidad Nueva', 'Desde', 'Hacia', 'Motivo', 'Referencia', 'Usuario'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos
    for row, movement in enumerate(movements, start=2):
        data = [
            movement.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            movement.get_movement_type_display(),
            str(movement.article),
            movement.quantity,
            movement.previous_quantity,
            movement.new_quantity,
            str(movement.from_location) if movement.from_location else '',
            str(movement.to_location) if movement.to_location else '',
            movement.reason,
            movement.reference,
            movement.user.username if movement.user else ''
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Ajustar anchos
    column_widths = [20, 15, 30, 12, 15, 15, 25, 25, 40, 20, 15]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=movimientos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


def export_loans_to_excel(loans):
    """Exportar préstamos a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Préstamos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'Artículo', 'Solicitante', 'ID/Matrícula', 'Contacto', 'Cantidad',
        'Fecha Préstamo', 'Fecha Devolución', 'Fecha Devolución Real', 
        'Estado', 'Días de Retraso', 'Aprobado Por', 'Notas'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Datos
    for row, loan in enumerate(loans, start=2):
        data = [
            str(loan.article),
            loan.borrower_name,
            loan.borrower_id,
            loan.borrower_contact,
            loan.quantity,
            loan.loan_date.strftime('%Y-%m-%d %H:%M:%S'),
            loan.due_date.strftime('%Y-%m-%d %H:%M:%S'),
            loan.return_date.strftime('%Y-%m-%d %H:%M:%S') if loan.return_date else '',
            loan.get_status_display(),
            loan.days_overdue if loan.is_overdue else 0,
            loan.approved_by.username if loan.approved_by else '',
            loan.notes
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            
            # Colorear filas de préstamos vencidos
            if loan.is_overdue:
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    
    # Ajustar anchos
    column_widths = [30, 25, 15, 20, 12, 20, 20, 20, 15, 15, 15, 40]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Crear respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=prestamos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response
